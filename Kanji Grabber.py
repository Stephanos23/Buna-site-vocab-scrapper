from selenium import webdriver
import openpyxl
import os
import time
wb_path = r"Kanji.xlsx"
#columns and rows for words
jpCol = 1 #column for japanese terms
defCol = 2 #column for definition
jpRow = 1 #row to use next for Japanese term
defRow = 1 #row to use next for English definition

wb = openpyxl.load_workbook(wb_path) #load workbook
ws = wb.worksheets[0] #Use first worksheet
kanjiSiteList = [] #Stores set of urls to grab Kanji from
kanjiLower = 121 #used to generate url for Kanji
kanjiUpper = 140 
for i in range(1,21): #generate url and add to list
    site = "https://buna.yorku.ca/japanese/ijlt/jp2000_kanji/jp2000_kanji_"+str(kanjiLower)+"-"+str(kanjiUpper)+".htm"
    #print(site)
    kanjiLower+=20
    kanjiUpper+=20
    kanjiSiteList.append(site)

#Selenium Driver
driver = webdriver.Chrome("C://WebDriver//bin//chromedriver.exe")
driver.get('https://buna.yorku.ca/japanese/imjp/ijlecnotes.html')

#Cleaning up text
def splitTextjp(text):
    global jpCol
    global jpRow
    list = text.splitlines() #splits text if they are on new line in list
    del list[1::2] #removes furigana or every even item on the list (1,3,5) list index start 0
    list.insert(1,' ') #spacing
    list.insert(3,' ') 
    resultString = ''.join(list) #concatenate items in list
    print(resultString) #prints resulting string
    ws.cell(row=jpRow, column=jpCol).value = resultString
    jpRow += 1
    print("-------------------") #seperate sets of Kanji
def splitTexten(text):
    global defCol
    global defRow
    temp = text.splitlines()
    for i in temp:
        print(i)
        ws.cell(row=defRow, column=defCol).value = i
        defRow += 1
        print("-------------------") #seperate sets of definitions
def Scrapper():
    global jpCol
    global jpRow
    global defCol
    global defRow
    for i in range(1,21): #checks tables from ranges 1-20
        for j in range(1,6): #loop 5 times
            try:
                text = driver.find_element_by_xpath("//body/div[1]/table["+str(i)+"]/tbody/tr[4]/td[2]/p["+str(j)+"]").text #grab core words (kanji)
                splitTextjp(text)
            except:
                print("No p["+str(j)+"]")

        try:
            text = driver.find_element_by_xpath("//body/div[1]/table["+str(i)+"]/tbody/tr[4]/td[3]").text #grabs core words (definition)
            splitTexten(text)
        except:
            print("No definition")

        for k in range(1,6): #loops 5 times
            try:
                text = driver.find_element_by_xpath("//body/div[1]/table["+str(i)+"]/tbody/tr[5]/td[2]/p["+str(k)+"]").text #grabs useful words (kanji)
                splitTextjp(text)
            except:
                print("No p["+str(k)+"]")

        try:
            text = driver.find_element_by_xpath("//body/div[1]/table["+str(i)+"]/tbody/tr[5]/td[3]").text #grabs useful words (definition)
            splitTexten(text)
        except:
            print("No definition")
        print("++++++++++++++++++++++++++++++++++++++++++") #seperator
    ws.cell(row=jpRow, column=jpCol).value = "---------------" #ends each set of kanji with line
    ws.cell(row=defRow, column=defCol).value = "---------------"
    jpRow += 1
    defRow += 1
    wb.save('Kanji.xlsx') #saves to workbook

for i in kanjiSiteList: #loads up each kanji page and grabs terms
    print(i)
    driver.get(i)
    time.sleep(5)
    Scrapper()
    time.sleep(5)
